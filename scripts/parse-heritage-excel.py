# -*- coding: utf-8 -*-
import json
import re
import zipfile
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(r"d:\电脑管家迁移文件\xwechat_files\wxid_nmn9our534th22_ed27\msg\file\2026-06\湖北省(副本) (4).xlsx")
OUT_PATH = Path(__file__).resolve().parent / "excel-scan.json"
VIDEOS_OUT = Path(__file__).resolve().parent.parent / "data" / "heritage-videos.js"

URL_RE = re.compile(r"https?://[^\s\"'<>]+", re.I)


def parse_sheet_hyperlinks(xlsx_path: Path):
    links_by_cell = {}
    rels = {}
    with zipfile.ZipFile(xlsx_path) as z:
        sheet_xml = "xl/worksheets/sheet1.xml"
        rels_xml = "xl/worksheets/_rels/sheet1.xml.rels"
        if sheet_xml in z.namelist():
            xml = z.read(sheet_xml).decode("utf-8")
            for m in re.finditer(r'<hyperlink[^>]*ref="([^"]+)"[^>]*r:id="([^"]+)"', xml):
                links_by_cell[m.group(1)] = m.group(2)
        if rels_xml in z.namelist():
            relxml = z.read(rels_xml).decode("utf-8")
            for m in re.finditer(r'Id="([^"]+)"[^>]*Target="([^"]+)"', relxml):
                rels[m.group(1)] = m.group(2)
    resolved = {}
    for cell_ref, rel_id in links_by_cell.items():
        target = rels.get(rel_id, rel_id)
        resolved[cell_ref] = target
    return resolved


def cell_ref_to_row_col(cell_ref: str):
    col_letters = re.match(r"([A-Z]+)", cell_ref).group(1)
    row = int(re.match(r"[A-Z]+(\d+)", cell_ref).group(1))
    col = 0
    for ch in col_letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return row, col


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["Sheet1"]
    hyperlinks = parse_sheet_hyperlinks(EXCEL_PATH)

    headers = [ws.cell(1, c).value for c in range(1, 12)]
    rows = []
    video_rows = []

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        if name.endswith("省") and len(name) <= 4:
            continue

        row_data = {
            "row": r,
            "name": name,
            "location": ws.cell(r, 2).value,
            "videoText": ws.cell(r, 9).value if ws.max_column >= 9 else None,
            "videoLinks": [],
        }

        for cell_ref, target in hyperlinks.items():
            hr, hc = cell_ref_to_row_col(cell_ref)
            if hr == r:
                row_data["videoLinks"].append({"col": hc, "url": target})

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if cell.hyperlink and cell.hyperlink.target:
                row_data["videoLinks"].append({"col": c, "url": cell.hyperlink.target})

        for field in ("videoText",):
            val = row_data.get(field)
            if isinstance(val, str):
                row_data["videoUrlsFromText"] = URL_RE.findall(val)

        if row_data["videoLinks"] or row_data.get("videoUrlsFromText"):
            video_rows.append(row_data)
        rows.append(row_data)

    result = {
        "headers": headers,
        "totalRows": len(rows),
        "videoRowCount": len(video_rows),
        "hyperlinksInXml": hyperlinks,
        "sampleRows": rows[:25],
        "videoRows": video_rows[:30],
    }
    OUT_PATH.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT_PATH}")
    print(f"headers={headers}")
    print(f"totalRows={len(rows)} videoRows={len(video_rows)} xmlLinks={len(hyperlinks)}")


if __name__ == "__main__":
    main()
