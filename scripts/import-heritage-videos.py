# -*- coding: utf-8 -*-
"""从 Excel 生成 data/heritage-videos.js"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = Path(
    r"d:\电脑管家迁移文件\xwechat_files\wxid_nmn9our534th22_ed27\msg\file\2026-06\湖北省(副本) (4).xlsx"
)
OUT_PATH = ROOT / "data" / "heritage-videos.js"
LIST_PATH = ROOT / "data" / "heritage-list.js"

URL_RE = re.compile(r"https?://[^\s\"'<>]+", re.I)
TITLE_RE = re.compile(r"【([^】]+)】")


def load_heritage_list():
    import subprocess
    script = (
        "const list=require('./data/heritage-list.js');"
        "console.log(JSON.stringify(list.map(x=>({name:x.name,slug:x.slug}))));"
    )
    result = subprocess.run(
        ["node", "-e", script],
        cwd=str(ROOT),
        capture_output=True,
        text=True,
        encoding="utf-8",
        check=True,
    )
    return json.loads(result.stdout.strip())


def normalize_name(name: str) -> str:
    if not name:
        return ""
    s = str(name).strip()
    s = re.sub(r"[（(].*?[)）]", "", s)
    s = re.sub(r"\s+", "", s)
    return s


def parse_video_cell(text: str):
    if not text or not str(text).strip():
        return None
    raw = str(text).strip()
    urls = URL_RE.findall(raw)
    if not urls:
        return None
    url = urls[0].rstrip(".,，。")
    title = URL_RE.sub("", raw).strip()
    title = re.sub(r"-哔哩哔哩", "", title)
    title = re.sub(r"^[【\s]+", "", title)
    title = re.sub(r"[】\s]+$", "", title)
    if not title or len(title) < 3:
        titles = TITLE_RE.findall(raw)
        title = max(titles, key=len).strip() if titles else "非遗影像"
    return {"title": title, "url": url}


def main():
    heritage_list = load_heritage_list()
    name_to_slug = {}
    for item in heritage_list:
        name_to_slug[normalize_name(item["name"])] = item["slug"]
        # prefix match keys for long names in excel vs list
        name_to_slug[item["name"].strip()] = item["slug"]

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet1"]

    by_slug = {}
    unmatched = []

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        if name.endswith("省") or name in ("北京", "天津", "上海", "重庆"):
            continue

        video = parse_video_cell(ws.cell(r, 9).value)
        if not video:
            continue

        norm = normalize_name(name)
        slug = name_to_slug.get(name) or name_to_slug.get(norm)
        if not slug:
            for item in heritage_list:
                ln = item["name"]
                if ln == name or ln.startswith(name) or name.startswith(ln):
                    slug = item["slug"]
                    break
                if normalize_name(ln) == norm:
                    slug = item["slug"]
                    break
        if not slug:
            unmatched.append(name)
            continue

        entry = {
            "title": video["title"],
            "url": video["url"],
            "source": "bilibili" if "b23.tv" in video["url"] or "bilibili" in video["url"] else "web",
        }
        by_slug[slug] = entry
        # 同名非遗（如两条「京剧」）共用同一视频
        for item in heritage_list:
            if item["name"] == name or normalize_name(item["name"]) == norm:
                by_slug[item["slug"]] = entry

    out = (
        "/**\n"
        " * 非遗影像链接（来源：湖北省(副本) Excel 视频超链接列）\n"
        f" * 共 {len(by_slug)} 条，按 heritage slug 索引\n"
        " */\n"
        "module.exports = "
        + json.dumps(by_slug, ensure_ascii=False, indent=2)
        + ";\n"
    )
    OUT_PATH.write_text(out, encoding="utf-8")
    print(f"Wrote {len(by_slug)} videos to {OUT_PATH}")
    if unmatched:
        print("Unmatched:", unmatched[:20], "..." if len(unmatched) > 20 else "")


if __name__ == "__main__":
    main()
